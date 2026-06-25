"""Reporting & export (FR-33..36). Reports run against local data (offline)
and every dataset can be returned as JSON, Excel (.xlsx) or PDF.
"""
from datetime import timedelta
from io import BytesIO

from django.db.models import Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view
from rest_framework.response import Response

from apps.billing.models import Invoice, InvoiceLine
from apps.customers.models import Prescription
from apps.inventory.models import Batch, Medicine


def _parse_dates(request):
    today = timezone.localdate()
    start = request.query_params.get('start') or (today - timedelta(days=30)).isoformat()
    end = request.query_params.get('end') or today.isoformat()
    return start, end


# --- Report builders: each returns (columns, rows, title) -------------------

def _stock_valuation():
    cols = ['Medicine', 'Manufacturer', 'Schedule', 'Qty', 'Cost', 'MRP', 'Stock value']
    rows = []
    for m in Medicine.objects.filter(is_active=True):
        for b in m.batches.filter(quantity__gt=0):
            rows.append([
                m.name, m.manufacturer, m.schedule, b.quantity,
                float(b.purchase_cost), float(b.mrp),
                round(b.quantity * float(b.purchase_cost), 2),
            ])
    return cols, rows, 'Stock & Valuation'


def _near_expiry(days=90):
    cutoff = timezone.localdate() + timedelta(days=days)
    cols = ['Medicine', 'Batch', 'Expiry', 'Days left', 'Qty', 'MRP']
    rows = [
        [b.medicine.name, b.batch_number, b.expiry_date.isoformat(),
         b.days_to_expiry, b.quantity, float(b.mrp)]
        for b in Batch.objects.filter(
            quantity__gt=0, expiry_date__gte=timezone.localdate(), expiry_date__lte=cutoff,
        ).select_related('medicine')
    ]
    return cols, rows, 'Near-Expiry'


def _low_stock():
    cols = ['Medicine', 'Stock', 'Reorder level', 'Reorder qty', 'Status']
    rows = [
        [m.name, m.total_stock, m.reorder_level, m.reorder_qty, m.stock_status]
        for m in Medicine.objects.filter(is_active=True)
        if m.stock_status in ('low_stock', 'out_of_stock')
    ]
    return cols, rows, 'Low / Out-of-Stock'


def _sales(start, end):
    cols = ['Invoice', 'Date', 'Customer', 'Payment', 'Taxable', 'CGST', 'SGST', 'Total']
    rows = []
    for inv in Invoice.objects.filter(
        created_at__date__gte=start, created_at__date__lte=end,
    ).select_related('customer'):
        rows.append([
            inv.number, inv.created_at.strftime('%Y-%m-%d'),
            inv.customer.name if inv.customer else 'Walk-in',
            inv.get_payment_mode_display(), float(inv.subtotal),
            float(inv.cgst), float(inv.sgst), float(inv.total),
        ])
    return cols, rows, 'Sales'


def _bills(start, end):
    """Every bill with line-level detail — one row per item sold (item 7)."""
    cols = ['Invoice', 'Date', 'Customer', 'Phone', 'Payment', 'Status', 'Item',
            'Batch', 'Expiry', 'Unit', 'Qty', 'Rate', 'GST%', 'Taxable', 'Tax', 'Amount']
    rows = []
    invoices = (Invoice.objects.filter(
        created_at__date__gte=start, created_at__date__lte=end)
        .select_related('customer').prefetch_related('lines').order_by('created_at'))
    for inv in invoices:
        cust = inv.customer
        for line in inv.lines.all():
            rows.append([
                inv.number, inv.created_at.strftime('%Y-%m-%d %H:%M'),
                cust.name if cust else 'Walk-in', cust.phone if cust else '',
                inv.get_payment_mode_display(), inv.get_status_display(),
                line.medicine_name, line.batch_number,
                line.expiry_date.strftime('%m/%Y') if line.expiry_date else '',
                'Loose' if line.unit_mode == 'loose' else 'Pack',
                line.quantity, float(line.rate), float(line.gst_rate),
                float(line.taxable_value), float(line.cgst_amount + line.sgst_amount),
                float(line.line_total),
            ])
    return cols, rows, 'Bills (detailed)'


def _gst_summary(start, end):
    cols = ['GST rate %', 'Taxable value', 'CGST', 'SGST', 'Total tax']
    buckets = {}
    for line in InvoiceLine.objects.filter(
        invoice__created_at__date__gte=start, invoice__created_at__date__lte=end,
    ):
        b = buckets.setdefault(float(line.gst_rate), [0.0, 0.0, 0.0])
        b[0] += float(line.taxable_value)
        b[1] += float(line.cgst_amount)
        b[2] += float(line.sgst_amount)
    rows = [[rate, round(v[0], 2), round(v[1], 2), round(v[2], 2), round(v[1] + v[2], 2)]
            for rate, v in sorted(buckets.items())]
    return cols, rows, 'GST Summary'


def _schedule_h1():
    cols = ['Date', 'Patient', 'Prescriber', 'Reg. no', 'Medicine', 'Qty', 'Invoice']
    rows = [
        [p.rx_date.isoformat(), p.patient_name, p.prescriber_name,
         p.prescriber_reg_no, p.medicine.name, p.quantity,
         p.invoice.number if p.invoice else '']
        for p in Prescription.objects.select_related('medicine', 'invoice')
        if p.medicine.schedule in ('H1', 'H', 'X')
    ]
    return cols, rows, 'Schedule H1 Register'


REPORTS = {
    'stock_valuation': lambda r: _stock_valuation(),
    'near_expiry': lambda r: _near_expiry(int(r.query_params.get('days', 90))),
    'low_stock': lambda r: _low_stock(),
    'sales': lambda r: _sales(*_parse_dates(r)),
    'bills': lambda r: _bills(*_parse_dates(r)),
    'gst_summary': lambda r: _gst_summary(*_parse_dates(r)),
    'schedule_h1': lambda r: _schedule_h1(),
}


@api_view(['GET'])
def report_view(request, key):
    builder = REPORTS.get(key)
    if not builder:
        return Response({'detail': f'Unknown report "{key}".'}, status=404)
    cols, rows, title = builder(request)
    # Note: 'format' is reserved by DRF content negotiation, so use 'export'.
    fmt = request.query_params.get('export', 'json')

    # Bills export is a two-sheet workbook: a per-bill summary + line items.
    if key == 'bills' and fmt == 'xlsx':
        return _bills_xlsx_response(*_parse_dates(request))

    if fmt == 'xlsx':
        return _xlsx_response(cols, rows, title)
    if fmt == 'pdf':
        return _pdf_response(cols, rows, title)
    return Response({
        'title': title, 'columns': cols, 'rows': rows, 'count': len(rows),
        'total': _numeric_total(cols, rows),
    })


def _numeric_total(cols, rows):
    """Sum the last column if it looks numeric — handy for sales/value totals."""
    try:
        return round(sum(float(r[-1]) for r in rows), 2)
    except (ValueError, TypeError, IndexError):
        return None


def _bills_xlsx_response(start, end):
    """Two-sheet bills workbook: 'Bills' (one row per invoice) + 'Bill items'
    (one row per item). Each bill is laid out separately so the shop can see
    totals at a glance and the full item detail on its own sheet."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    head_fill = PatternFill('solid', fgColor='1E2640')

    def style_header(ws):
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = head_fill
        for i in range(1, ws.max_column + 1):
            col = ws.cell(row=1, column=i)
            ws.column_dimensions[col.column_letter].width = max(12, len(str(col.value)) + 3)

    invoices = (Invoice.objects.filter(
        created_at__date__gte=start, created_at__date__lte=end)
        .select_related('customer').prefetch_related('lines').order_by('created_at'))

    # Sheet 1 — bill summary
    s1 = wb.active
    s1.title = 'Bills'
    s1.append(['Invoice', 'Date', 'Customer', 'Phone', 'Payment', 'Status',
               'Items', 'Taxable', 'CGST', 'SGST', 'Discount', 'Total'])
    for inv in invoices:
        c = inv.customer
        s1.append([
            inv.number, inv.created_at.strftime('%Y-%m-%d %H:%M'),
            c.name if c else 'Walk-in', c.phone if c else '',
            inv.get_payment_mode_display(), inv.get_status_display(),
            inv.lines.count(), float(inv.subtotal), float(inv.cgst),
            float(inv.sgst), float(inv.discount), float(inv.total),
        ])
    style_header(s1)

    # Sheet 2 — line items
    s2 = wb.create_sheet('Bill items')
    s2.append(['Invoice', 'Date', 'Item', 'Batch', 'Expiry', 'Unit', 'Qty',
               'Rate', 'GST%', 'Taxable', 'Tax', 'Amount'])
    for inv in invoices:
        for line in inv.lines.all():
            s2.append([
                inv.number, inv.created_at.strftime('%Y-%m-%d'),
                line.medicine_name, line.batch_number,
                line.expiry_date.strftime('%m/%Y') if line.expiry_date else '',
                'Loose' if line.unit_mode == 'loose' else 'Pack', line.quantity,
                float(line.rate), float(line.gst_rate), float(line.taxable_value),
                float(line.cgst_amount + line.sgst_amount), float(line.line_total),
            ])
    style_header(s2)

    buf = BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="Bills_{start}_to_{end}.xlsx"'
    return resp


def _xlsx_response(cols, rows, title):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append(cols)
    header_fill = PatternFill('solid', fgColor='1E2640')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
    for row in rows:
        ws.append(row)
    for i, col in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, len(str(col)) + 4)

    buf = BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}.xlsx"'
    return resp


def _pdf_response(cols, rows, title):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    data = [cols] + [[str(c) for c in row] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e2640')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cfd6e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f7fc')]),
    ]))
    elems = [
        Paragraph(f'<b>{title}</b>', styles['Title']),
        Paragraph(f'Generated {timezone.localtime():%Y-%m-%d %H:%M} · {len(rows)} rows',
                  styles['Normal']),
        Spacer(1, 4 * mm), table,
    ]
    doc.build(elems)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{title.replace(" ", "_")}.pdf"'
    return resp
